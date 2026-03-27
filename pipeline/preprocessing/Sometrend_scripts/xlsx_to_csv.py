import argparse
import csv
import math
import re
import sys
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path


INVALID_FILENAME_CHARS = r'[<>:"/\\|?*\x00-\x1f]'
FORMULA_PREFIXES = ("=", "+", "-", "@")
DATE_LIKE_PATTERNS = (
    re.compile(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}$"),
    re.compile(r"^\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}$"),
    re.compile(r"^\d{4}[-/.]\d{1,2}$"),
    re.compile(r"^\d{1,2}:\d{2}(?::\d{2})?$"),
)
NUMERIC_LIKE_PATTERNS = (
    re.compile(r"^[+-]?\d+$"),
    re.compile(r"^[+-]?\d+\.\d+$"),
    re.compile(r"^[+-]?(?:\d+(?:\.\d+)?|\.\d+)[Ee][+-]?\d+$"),
)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert one or more XLSX files to CSV files."
    )
    parser.add_argument(
        "inputs",
        nargs="+",
        help="One or more .xlsx/.xlsm files to convert.",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        help="Optional directory where CSV files will be written.",
    )
    parser.add_argument(
        "--encoding",
        default="utf-8-sig",
        help="CSV encoding. Default: utf-8-sig",
    )
    parser.add_argument(
        "--replace-original",
        action="store_true",
        help="Remove the original XLSX/XLSM file after a successful conversion.",
    )
    return parser


def safe_name(value: str) -> str:
    cleaned = re.sub(INVALID_FILENAME_CHARS, "_", value).strip()
    return cleaned or "sheet"


def normalize_input_path(raw_path: str) -> Path:
    return Path(raw_path.strip().strip('"')).expanduser().resolve()


def zero_pad_width(number_format: str) -> int | None:
    primary_format = (number_format or "").split(";")[0].strip()
    if re.fullmatch(r"0+", primary_format):
        return len(primary_format)
    return None


def decimal_to_text(value: Decimal) -> str:
    text = format(value, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def looks_like_excel_auto_conversion_target(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return False
    if stripped.startswith(FORMULA_PREFIXES):
        return True
    if any(pattern.fullmatch(stripped) for pattern in DATE_LIKE_PATTERNS):
        return True
    return any(pattern.fullmatch(stripped) for pattern in NUMERIC_LIKE_PATTERNS)


def protect_for_excel_text(text: str, force: bool = False) -> str:
    if not text or text.startswith("\t"):
        return text
    if force or looks_like_excel_auto_conversion_target(text):
        return f"\t{text}"
    return text


def apply_simple_number_format(cell, text: str, is_integer: bool) -> str:
    if not is_integer:
        return text

    width = zero_pad_width(getattr(cell, "number_format", ""))
    if width is None:
        return text

    sign = ""
    digits = text
    if text.startswith("-"):
        sign = "-"
        digits = text[1:]

    if digits.isdigit():
        return f"{sign}{digits.zfill(width)}"
    return text


def format_datetime_value(value) -> str:
    if isinstance(value, datetime):
        if value.microsecond:
            return value.isoformat(sep=" ")
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        if value.microsecond:
            return value.isoformat(timespec="microseconds")
        return value.isoformat(timespec="seconds")
    return str(value)


def cell_to_csv_value(cell) -> str:
    value = getattr(cell, "value", None)

    if value is None:
        return ""
    if isinstance(value, str):
        return protect_for_excel_text(value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date, time)):
        return protect_for_excel_text(format_datetime_value(value), force=True)
    if isinstance(value, int):
        text = apply_simple_number_format(cell, str(value), is_integer=True)
        return protect_for_excel_text(text, force=True)
    if isinstance(value, Decimal):
        is_integer = value == value.to_integral_value()
        text = str(int(value)) if is_integer else decimal_to_text(value)
        text = apply_simple_number_format(cell, text, is_integer=is_integer)
        return protect_for_excel_text(text, force=True)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return str(value)
        if value.is_integer():
            text = apply_simple_number_format(cell, str(int(value)), is_integer=True)
            return protect_for_excel_text(text, force=True)
        try:
            text = decimal_to_text(Decimal(str(value)))
        except InvalidOperation:
            text = str(value)
        text = apply_simple_number_format(cell, text, is_integer=False)
        return protect_for_excel_text(text, force=True)

    return protect_for_excel_text(str(value))


def load_workbook_safe():
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "openpyxl is required. Install it with: pip install openpyxl",
            file=sys.stderr,
        )
        raise SystemExit(2)
    return load_workbook


def csv_path_for_sheet(xlsx_path: Path, sheet_name: str, sheet_count: int, output_dir: Path) -> Path:
    if sheet_count == 1:
        filename = f"{xlsx_path.stem}.csv"
    else:
        filename = f"{xlsx_path.stem}__{safe_name(sheet_name)}.csv"
    return output_dir / filename


def convert_file(
    xlsx_path: Path,
    output_dir: Path,
    encoding: str,
    replace_original: bool,
) -> int:
    if not xlsx_path.exists():
        print(f"[ERROR] File not found: {xlsx_path}", file=sys.stderr)
        return 1

    if xlsx_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        print(f"[ERROR] Unsupported file type: {xlsx_path}", file=sys.stderr)
        return 1

    load_workbook = load_workbook_safe()

    try:
        workbook = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
    except Exception as exc:
        print(f"[ERROR] Failed to open workbook: {xlsx_path}", file=sys.stderr)
        print(f"        {exc}", file=sys.stderr)
        return 1

    try:
        sheet_names = workbook.sheetnames
        output_dir.mkdir(parents=True, exist_ok=True)

        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            csv_path = csv_path_for_sheet(
                xlsx_path=xlsx_path,
                sheet_name=sheet_name,
                sheet_count=len(sheet_names),
                output_dir=output_dir,
            )

            with csv_path.open("w", newline="", encoding=encoding) as csv_file:
                writer = csv.writer(csv_file)
                for row in worksheet.iter_rows():
                    writer.writerow([cell_to_csv_value(cell) for cell in row])

            print(f"[OK] {xlsx_path.name} -> {csv_path}")
    finally:
        workbook.close()

    if replace_original:
        try:
            xlsx_path.unlink()
            print(f"[OK] Removed original file: {xlsx_path}")
        except Exception as exc:
            print(f"[ERROR] Failed to remove original file: {xlsx_path}", file=sys.stderr)
            print(f"        {exc}", file=sys.stderr)
            return 1

    return 0


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    if args.output_dir:
        common_output_dir = Path(args.output_dir).expanduser().resolve()
    else:
        common_output_dir = None

    exit_code = 0
    for raw_input in args.inputs:
        xlsx_path = normalize_input_path(raw_input)
        output_dir = common_output_dir or xlsx_path.parent
        result = convert_file(
            xlsx_path=xlsx_path,
            output_dir=output_dir,
            encoding=args.encoding,
            replace_original=args.replace_original,
        )
        if result != 0:
            exit_code = result

    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())
